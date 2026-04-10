#!/usr/bin/env python3
"""
Generate a multi-level trace chain report.

Walks the full derivation hierarchy:
  BRQ -> BR -> CTRL -> FR/NFR -> US -> TASK -> TEST

Reports complete chains (traced end-to-end) and broken chains
(where a link is missing at some level).

Usage:
    python scripts/generate-trace-chains.py [--path PATH] [--output PATH] [--format md|both]

Requires: pip install pyyaml openpyxl
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import date
from collections import defaultdict

try:
    import yaml
except ImportError:
    print("ERROR: pip install pyyaml")
    sys.exit(1)


FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---", re.DOTALL)
WIKILINK_RE = re.compile(r"\[\[([^\]|]+)(?:\|[^\]]+)?\]\]")

FORWARD_FIELDS = [
    "derives_from", "depends_on", "parent_epic", "parent_overview",
    "implements", "implements_control", "part_of_story",
    "delivers", "verifies", "verifies_control", "covers_criteria",
    "affects", "persona", "source_docs", "source_ref",
    "requirements_included", "epics_included", "related_epics",
    "related_adrs", "blocks", "implemented_by", "verified_by",
    "related_requirements", "derived_requirements", "derived_controls",
    "delivered_by", "related_brqs", "related_ctrls",
]

# The hierarchy levels (top to bottom).
# Each level defines: prefix, what field on CHILD points to this level
HIERARCHY = [
    {"prefix": "BRQ",  "label": "Business Requirement"},
    {"prefix": "BR",   "label": "Business Rule"},
    {"prefix": "CTRL", "label": "Control"},
    {"prefix": "CON",  "label": "Constraint"},
    {"prefix": "EPIC", "label": "Epic"},
    {"prefix": "FR",   "label": "Functional Requirement"},
    {"prefix": "NFR",  "label": "Non-Functional Requirement"},
    {"prefix": "US",   "label": "User Story"},
    {"prefix": "TASK", "label": "Task"},
    {"prefix": "TEST", "label": "Acceptance Test"},
]

# How child artifacts reference parents (child_prefix -> {field -> parent_prefix})
UPWARD_LINKS = {
    "BR":   {"derives_from": "BRQ"},
    "CTRL": {"derives_from": "BRQ", "implements_control": "BR"},
    "CON":  {"derives_from": "BRQ"},
    "EPIC": {"derives_from": "BRQ", "source_docs": "VISION"},
    "FR":   {"parent_epic": "EPIC", "derives_from": "BRQ"},
    "NFR":  {"parent_epic": "EPIC", "derives_from": "BRQ"},
    "US":   {"parent_epic": "EPIC", "delivers": "FR"},
    "TASK": {"implements": "FR", "part_of_story": "US"},
    "TEST": {"verifies": "FR"},
}

# Reverse inference: what links DOWN from a parent
# parent_prefix -> [(child_prefix, field_on_child)]
DOWNWARD_MAP = defaultdict(list)
for child_prefix, fields in UPWARD_LINKS.items():
    for field, parent_prefix in fields.items():
        DOWNWARD_MAP[parent_prefix].append((child_prefix, field))


def extract_frontmatter(filepath: Path) -> dict | None:
    text = filepath.read_text(encoding="utf-8")
    match = FRONTMATTER_RE.match(text)
    if not match:
        return None
    try:
        return yaml.safe_load(match.group(1))
    except yaml.YAMLError:
        return None


def extract_links(value) -> list[str]:
    links = []
    if isinstance(value, str):
        for m in WIKILINK_RE.finditer(value):
            links.append(m.group(1))
        if not links and value.strip():
            links.append(value.strip())
    elif isinstance(value, list):
        for item in value:
            links.extend(extract_links(item))
    return links


def get_prefix(artifact_id: str) -> str:
    return artifact_id.split("-")[0]


def build_artifacts(root: Path) -> dict:
    artifacts = {}
    for md_file in sorted(root.rglob("*.md")):
        rel = md_file.relative_to(root)
        if any(part.startswith("_") for part in rel.parts) and "examples" not in str(rel).lower():
            continue
        if "templates" in rel.parts:
            continue

        fm = extract_frontmatter(md_file)
        if fm is None or "id" not in fm:
            continue

        aid = fm["id"]
        artifacts[aid] = {
            "title": fm.get("title", ""),
            "status": fm.get("status", ""),
            "priority": fm.get("priority", ""),
            "domain": fm.get("domain", ""),
            "_links": {},
        }
        for field in FORWARD_FIELDS:
            if field in fm and fm[field]:
                linked = extract_links(fm[field])
                if linked:
                    artifacts[aid]["_links"][field] = linked

    return artifacts


def build_children_index(artifacts: dict) -> dict:
    """Build parent_id -> [(child_id, field)] index, deduplicated by (parent, child) pair."""
    children = defaultdict(list)
    seen_edges = set()  # (parent_id, child_id) pairs to avoid duplicates

    for aid, info in artifacts.items():
        prefix = get_prefix(aid)
        if prefix not in UPWARD_LINKS:
            continue
        for field, parent_prefix in UPWARD_LINKS[prefix].items():
            for target in info["_links"].get(field, []):
                if get_prefix(target) == parent_prefix:
                    edge = (target, aid)
                    if edge not in seen_edges:
                        seen_edges.add(edge)
                        children[target].append((aid, field))

    # Also handle legacy reverse fields (only if edge not already registered)
    for aid, info in artifacts.items():
        for rev_field, child_prefix in [("delivered_by", "US"), ("implemented_by", "TASK"), ("verified_by", "TEST")]:
            for target in info["_links"].get(rev_field, []):
                if get_prefix(target) == child_prefix:
                    edge = (aid, target)
                    if edge not in seen_edges:
                        seen_edges.add(edge)
                        children[aid].append((target, rev_field))

    return children


def walk_chains_down(start_id: str, artifacts: dict, children_index: dict, chain: list, all_chains: list):
    """Recursively walk down from start_id, collecting all leaf chains."""
    kids = children_index.get(start_id, [])
    if not kids:
        all_chains.append(list(chain))
        return

    for child_id, field in kids:
        if child_id not in artifacts:
            chain.append({"id": child_id, "missing": True, "field": field})
            all_chains.append(list(chain))
            chain.pop()
            continue
        chain.append({"id": child_id, "missing": False, "field": field})
        walk_chains_down(child_id, artifacts, children_index, chain, all_chains)
        chain.pop()


def analyze_chains(artifacts: dict) -> dict:
    """Build trace chains starting from top-level artifacts (BRQ, EPIC, VISION)."""
    children_index = build_children_index(artifacts)

    # Find root artifacts: BRQ and standalone EPICs (no derives_from BRQ)
    roots = []
    for aid in sorted(artifacts):
        prefix = get_prefix(aid)
        if prefix == "BRQ":
            roots.append(aid)
        elif prefix == "EPIC":
            # Include EPIC as root if it doesn't derive from a BRQ
            links = artifacts[aid]["_links"]
            derives = links.get("derives_from", [])
            has_brq_parent = any(get_prefix(d) == "BRQ" for d in derives)
            if not has_brq_parent:
                roots.append(aid)

    # Also find FR/NFR that have no parent epic (direct roots)
    for aid in sorted(artifacts):
        prefix = get_prefix(aid)
        if prefix in ("FR", "NFR"):
            links = artifacts[aid]["_links"]
            if not links.get("parent_epic") and not links.get("derives_from"):
                roots.append(aid)

    all_chains = []
    for root_id in roots:
        chain = [{"id": root_id, "missing": False, "field": "root"}]
        walk_chains_down(root_id, artifacts, children_index, chain, all_chains)

    return {
        "roots": roots,
        "chains": all_chains,
        "children_index": children_index,
    }


def classify_chain(chain: list) -> str:
    """Classify a chain as complete, partial, or broken."""
    prefixes = [get_prefix(step["id"]) for step in chain]
    has_missing = any(step.get("missing") for step in chain)

    if has_missing:
        return "broken"

    # A complete chain should reach TEST (or at least TASK)
    if "TEST" in prefixes:
        return "complete"
    elif "TASK" in prefixes or "US" in prefixes:
        return "partial"
    else:
        return "stub"


def generate_markdown(artifacts: dict, analysis: dict) -> str:
    chains = analysis["chains"]

    classified = {"complete": [], "partial": [], "stub": [], "broken": []}
    for chain in chains:
        cls = classify_chain(chain)
        classified[cls].append(chain)

    lines = [
        "---",
        "id: META-TRACE-CHAINS",
        "title: Multi-Level Trace Chain Report",
        f"updated: {date.today().isoformat()}",
        "---",
        "",
        "# Trace Chain Report",
        "",
        "> Auto-generated by `scripts/generate-trace-chains.py`.",
        "> Do not edit manually.",
        "",
        "## Summary",
        "",
        f"| Metric | Count |",
        f"|--------|-------|",
        f"| Root artifacts | {len(analysis['roots'])} |",
        f"| Total chains | {len(chains)} |",
        f"| Complete (reaches TEST) | {len(classified['complete'])} |",
        f"| Partial (reaches US/TASK, no TEST) | {len(classified['partial'])} |",
        f"| Stub (stops at FR/NFR or above) | {len(classified['stub'])} |",
        f"| Broken (missing artifact) | {len(classified['broken'])} |",
        "",
    ]

    if chains:
        total = len(chains)
        complete_pct = (len(classified["complete"]) / total) * 100
        lines.append(f"**End-to-end coverage: {complete_pct:.1f}%**")
        lines.append("")

    # Broken chains first (most actionable)
    if classified["broken"]:
        lines.append("## Broken Chains")
        lines.append("")
        lines.append("These chains reference artifacts that do not exist in the vault:")
        lines.append("")
        for chain in classified["broken"]:
            chain_str = format_chain(chain)
            lines.append(f"- {chain_str}")
        lines.append("")

    # Stubs (requirements not yet decomposed)
    if classified["stub"]:
        lines.append("## Stub Chains")
        lines.append("")
        lines.append("These chains stop at requirements level with no delivery or verification:")
        lines.append("")
        for chain in classified["stub"]:
            chain_str = format_chain(chain)
            lines.append(f"- {chain_str}")
        lines.append("")

    # Partial
    if classified["partial"]:
        lines.append("## Partial Chains (no test coverage)")
        lines.append("")
        for chain in classified["partial"]:
            chain_str = format_chain(chain)
            lines.append(f"- {chain_str}")
        lines.append("")

    # Complete
    if classified["complete"]:
        lines.append("## Complete Chains")
        lines.append("")
        for chain in classified["complete"]:
            chain_str = format_chain(chain)
            lines.append(f"- {chain_str}")
        lines.append("")

    # Unreachable artifacts (not part of any chain)
    all_in_chains = set()
    for chain in chains:
        for step in chain:
            all_in_chains.add(step["id"])

    unreachable = sorted(set(artifacts.keys()) - all_in_chains
                         - {aid for aid in artifacts if get_prefix(aid) in ("VISION", "META", "GLOSS", "CODEMAP", "PERSONA", "JOURNEY", "ASSUM", "ADR", "CR", "REL", "DM", "ARCH", "CONTRACT", "SRC")})
    if unreachable:
        lines.append("## Unreachable Artifacts")
        lines.append("")
        lines.append("These artifacts are not part of any derivation chain:")
        lines.append("")
        for uid in unreachable:
            info = artifacts.get(uid, {})
            lines.append(f"- [[{uid}]] ({get_prefix(uid)}) — {info.get('title', '')}")
        lines.append("")

    return "\n".join(lines)


def format_chain(chain: list) -> str:
    parts = []
    for step in chain:
        aid = step["id"]
        if step.get("missing"):
            parts.append(f"~~[[{aid}]]~~ (MISSING)")
        else:
            parts.append(f"[[{aid}]]")
    return " -> ".join(parts)


def generate_xlsx(artifacts: dict, analysis: dict, output_path: Path):
    """Generate an XLSX version of the trace chain report."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("WARNING: openpyxl not installed, skipping XLSX output")
        return

    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    COMPLETE_FILL = PatternFill("solid", fgColor="C6EFCE")
    PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
    STUB_FILL = PatternFill("solid", fgColor="D9D9D9")
    BROKEN_FILL = PatternFill("solid", fgColor="FFC7CE")
    THIN_BORDER = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Trace Chains"

    # Find max chain length for column count
    max_len = max((len(c) for c in analysis["chains"]), default=1)

    # Headers
    headers = ["Chain #", "Classification", "Depth"]
    for i in range(max_len):
        headers.append(f"Level {i}")
    headers.append("Gap Description")

    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    fill_map = {
        "complete": COMPLETE_FILL,
        "partial": PARTIAL_FILL,
        "stub": STUB_FILL,
        "broken": BROKEN_FILL,
    }

    for idx, chain in enumerate(analysis["chains"]):
        r = 2 + idx
        cls = classify_chain(chain)

        ws.cell(row=r, column=1, value=idx + 1).font = BODY_FONT
        cls_cell = ws.cell(row=r, column=2, value=cls.upper())
        cls_cell.font = Font(bold=True, name="Arial", size=10)
        cls_cell.fill = fill_map.get(cls, STUB_FILL)
        ws.cell(row=r, column=3, value=len(chain)).font = BODY_FONT

        gap_parts = []
        for level_idx, step in enumerate(chain):
            cell = ws.cell(row=r, column=4 + level_idx, value=step["id"])
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if step.get("missing"):
                cell.fill = BROKEN_FILL
                cell.font = Font(name="Arial", size=10, strikethrough=True, color="FF0000")
                gap_parts.append(f"{step['id']} missing")

        if cls == "partial":
            prefixes = {get_prefix(s["id"]) for s in chain}
            if "TEST" not in prefixes:
                gap_parts.append("No TEST artifact")
        elif cls == "stub":
            gap_parts.append("Stops before delivery/verification")

        ws.cell(row=r, column=4 + max_len, value="; ".join(gap_parts)).font = BODY_FONT

        for c in range(1, 4 + max_len + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8
    for i in range(max_len):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(4 + i)].width = 20
    ws.column_dimensions[get_column_letter(4 + max_len)].width = 35

    ws.freeze_panes = ws.cell(row=2, column=4)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Trace Chain Summary").font = Font(bold=True, name="Arial", size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}").font = Font(italic=True, name="Arial", size=10)

    chains = analysis["chains"]
    classified = {"complete": 0, "partial": 0, "stub": 0, "broken": 0}
    for chain in chains:
        classified[classify_chain(chain)] += 1

    summary_data = [
        ("Root Artifacts", len(analysis["roots"])),
        ("Total Chains", len(chains)),
        ("Complete (reaches TEST)", classified["complete"]),
        ("Partial (no TEST)", classified["partial"]),
        ("Stub (no delivery)", classified["stub"]),
        ("Broken (missing ref)", classified["broken"]),
    ]

    fills = [None, None, COMPLETE_FILL, PARTIAL_FILL, STUB_FILL, BROKEN_FILL]

    for i, (label, count) in enumerate(summary_data):
        r = 4 + i
        ws2.cell(row=r, column=1, value=label).font = BODY_FONT
        c = ws2.cell(row=r, column=2, value=count)
        c.font = Font(bold=True, name="Arial", size=10)
        if fills[i]:
            c.fill = fills[i]
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER

    if chains:
        total = len(chains)
        pct = (classified["complete"] / total) * 100
        ws2.cell(row=11, column=1, value="End-to-end Coverage").font = Font(bold=True, name="Arial", size=12)
        ws2.cell(row=11, column=2, value=f"{pct:.1f}%").font = Font(bold=True, name="Arial", size=12)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    wb.save(str(output_path))
    print(f"Trace chain XLSX saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Generate multi-level trace chain report")
    parser.add_argument("--path", default=".", help="Root of requirements vault")
    parser.add_argument("--output", default=None, help="Output file (default: 05-quality/traceability/)")
    parser.add_argument("--format", default="both", choices=["md", "both"], help="Output format")
    args = parser.parse_args()

    root = Path(args.path)
    artifacts = build_artifacts(root)
    analysis = analyze_chains(artifacts)

    print(f"Found {len(artifacts)} artifacts, {len(analysis['roots'])} roots, {len(analysis['chains'])} chains")

    output_dir = Path(args.output) if args.output else root / "05-quality" / "traceability"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Markdown report
    md_content = generate_markdown(artifacts, analysis)
    md_path = output_dir / "TRACE-CHAINS.md"
    md_path.write_text(md_content, encoding="utf-8")
    print(f"Trace chain report saved: {md_path}")

    # XLSX
    if args.format == "both":
        xlsx_path = output_dir / "TRACE-CHAINS.xlsx"
        generate_xlsx(artifacts, analysis, xlsx_path)


if __name__ == "__main__":
    main()
