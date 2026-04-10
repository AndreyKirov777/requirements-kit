#!/usr/bin/env python3
"""
Generate a traceability matrix as an XLSX spreadsheet.

Produces a cross-reference matrix with requirements (FR/NFR) as rows and
verification/delivery artifacts (US, TASK, TEST) as columns. Cells show
the link relationship. A coverage summary sheet highlights gaps.

Usage:
    python scripts/generate-traceability-matrix.py [--path PATH] [--output PATH]

Requires: pip install pyyaml openpyxl
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import date

try:
    import yaml
except ImportError:
    print("ERROR: pip install pyyaml")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)


FRONTMATTER_RE = re.compile(r"^---\s*\n(.*?)\n---", re.DOTALL)
WIKILINK_RE = re.compile(r"\[\[([^\]|]+)(?:\|[^\]]+)?\]\]")

# Forward link fields (child -> parent)
FORWARD_FIELDS = [
    "derives_from", "depends_on", "parent_epic", "parent_overview",
    "implements", "implements_control", "part_of_story",
    "delivers", "verifies", "verifies_control", "covers_criteria",
    "affects", "persona", "source_docs", "source_ref",
    "requirements_included", "epics_included", "related_epics",
    "related_adrs", "blocks", "implemented_by", "verified_by",
    "related_requirements", "derived_requirements", "derived_controls",
    "delivered_by", "related_brqs", "related_ctrls",
]

# The derivation hierarchy (top to bottom)
LAYER_ORDER = ["BRQ", "BR", "CTRL", "CON", "EPIC", "FR", "NFR", "US", "TASK", "TEST"]

# Matrix row types (requirements to trace)
ROW_TYPES = {"FR", "NFR"}

# Matrix column types (delivery/verification artifacts)
COL_TYPES = {"US", "TASK", "TEST"}

# Relationship fields: which field on the COLUMN artifact points to the ROW artifact
REL_MAP = {
    "US":   {"field": "delivers",   "label": "delivers"},
    "TASK": {"field": "implements", "label": "implements"},
    "TEST": {"field": "verifies",   "label": "verifies"},
}

# Colors
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="2F5496")
GAP_FILL = PatternFill("solid", fgColor="FFC7CE")  # red-ish for gaps
OK_FILL = PatternFill("solid", fgColor="C6EFCE")    # green for covered
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")  # yellow for partial
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def extract_frontmatter(filepath: Path) -> dict | None:
    text = filepath.read_text(encoding="utf-8")
    match = FRONTMATTER_RE.match(text)
    if not match:
        return None
    try:
        return yaml.safe_load(match.group(1))
    except yaml.YAMLError:
        return None


def extract_links(value) -> list[str]:
    links = []
    if isinstance(value, str):
        for m in WIKILINK_RE.finditer(value):
            links.append(m.group(1))
        if not links and value.strip():
            links.append(value.strip())
    elif isinstance(value, list):
        for item in value:
            links.extend(extract_links(item))
    return links


def get_prefix(artifact_id: str) -> str:
    return artifact_id.split("-")[0]


def build_artifacts(root: Path) -> dict:
    """Build artifact map: id -> {frontmatter fields + _links}."""
    artifacts = {}
    for md_file in sorted(root.rglob("*.md")):
        rel = md_file.relative_to(root)
        if any(part.startswith("_") for part in rel.parts) and "examples" not in str(rel).lower():
            continue
        if "templates" in rel.parts:
            continue

        fm = extract_frontmatter(md_file)
        if fm is None or "id" not in fm:
            continue

        aid = fm["id"]
        artifacts[aid] = {
            "title": fm.get("title", ""),
            "status": fm.get("status", ""),
            "priority": fm.get("priority", ""),
            "domain": fm.get("domain", ""),
            "owner": fm.get("owner", ""),
            "_links": {},
        }

        for field in FORWARD_FIELDS:
            if field in fm and fm[field]:
                linked = extract_links(fm[field])
                if linked:
                    artifacts[aid]["_links"][field] = linked

    return artifacts


def compute_reverse_links(artifacts: dict) -> dict:
    """Build reverse index: target_id -> {source_id: [field_names]}."""
    reverse = {}
    for aid, info in artifacts.items():
        for field, targets in info["_links"].items():
            for target in targets:
                reverse.setdefault(target, {}).setdefault(aid, []).append(field)
    return reverse


def write_matrix_sheet(ws, artifacts, reverse):
    """Write the cross-reference matrix sheet."""
    # Collect row and column artifacts
    row_ids = sorted([aid for aid in artifacts if get_prefix(aid) in ROW_TYPES],
                     key=lambda x: (LAYER_ORDER.index(get_prefix(x)) if get_prefix(x) in LAYER_ORDER else 99, x))
    col_ids_by_type = {}
    for ctype in ["US", "TASK", "TEST"]:
        col_ids_by_type[ctype] = sorted([aid for aid in artifacts if get_prefix(aid) == ctype])

    all_col_ids = []
    col_type_ranges = {}
    for ctype in ["US", "TASK", "TEST"]:
        start = len(all_col_ids)
        all_col_ids.extend(col_ids_by_type[ctype])
        col_type_ranges[ctype] = (start, len(all_col_ids))

    if not row_ids:
        ws["A1"] = "No FR/NFR artifacts found."
        return

    # Fixed columns: ID, Title, Status, Priority, Domain
    fixed_cols = 5
    header_row = 2  # row 1 for group headers

    # Group header row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=fixed_cols)
    cell = ws.cell(row=1, column=1, value="Requirement")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

    for ctype in ["US", "TASK", "TEST"]:
        start_idx, end_idx = col_type_ranges[ctype]
        if start_idx == end_idx:
            continue
        start_col = fixed_cols + 1 + start_idx
        end_col = fixed_cols + end_idx
        if start_col <= end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=f"{ctype} ({REL_MAP[ctype]['label']})")
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    # Coverage summary column
    summary_col = fixed_cols + len(all_col_ids) + 1
    ws.cell(row=1, column=summary_col, value="Coverage").font = HEADER_FONT
    ws.cell(row=1, column=summary_col).fill = HEADER_FILL
    ws.cell(row=1, column=summary_col).alignment = Alignment(horizontal="center")

    # Sub-headers
    sub_headers = ["ID", "Title", "Status", "Priority", "Domain"]
    for i, h in enumerate(sub_headers):
        cell = ws.cell(row=header_row, column=i + 1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    for i, cid in enumerate(all_col_ids):
        cell = ws.cell(row=header_row, column=fixed_cols + 1 + i, value=cid)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(text_rotation=90, horizontal="center")

    cell = ws.cell(row=header_row, column=summary_col, value="Status")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.border = THIN_BORDER

    # Data rows
    for row_idx, rid in enumerate(row_ids):
        r = header_row + 1 + row_idx
        info = artifacts[rid]

        ws.cell(row=r, column=1, value=rid).font = BODY_FONT
        ws.cell(row=r, column=2, value=info["title"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=info["status"]).font = BODY_FONT
        ws.cell(row=r, column=4, value=info["priority"]).font = BODY_FONT
        ws.cell(row=r, column=5, value=info["domain"]).font = BODY_FONT

        for c in range(1, fixed_cols + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

        has_us = False
        has_task = False
        has_test = False

        for col_idx, cid in enumerate(all_col_ids):
            cell = ws.cell(row=r, column=fixed_cols + 1 + col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # Check if cid links to rid
            col_links = artifacts.get(cid, {}).get("_links", {})
            linked = False
            link_field = ""
            cprefix = get_prefix(cid)

            if cprefix in REL_MAP:
                field = REL_MAP[cprefix]["field"]
                targets = col_links.get(field, [])
                if rid in targets:
                    linked = True
                    link_field = field

            # Also check reverse: if rid has delivered_by/implemented_by/verified_by pointing to cid
            rid_links = artifacts[rid].get("_links", {})
            for rev_field in ["delivered_by", "implemented_by", "verified_by"]:
                if cid in rid_links.get(rev_field, []):
                    linked = True
                    link_field = rev_field

            if linked:
                cell.value = link_field
                cell.font = LINK_FONT
                if cprefix == "US":
                    has_us = True
                elif cprefix == "TASK":
                    has_task = True
                elif cprefix == "TEST":
                    has_test = True

        # Coverage summary
        summary_cell = ws.cell(row=r, column=summary_col)
        summary_cell.border = THIN_BORDER
        summary_cell.alignment = Alignment(horizontal="center")
        summary_cell.font = Font(bold=True, name="Arial", size=10)

        if has_us and has_test:
            summary_cell.value = "Covered"
            summary_cell.fill = OK_FILL
        elif has_us or has_task or has_test:
            parts = []
            if not has_us and not has_task:
                parts.append("no delivery")
            if not has_test:
                parts.append("no test")
            summary_cell.value = "Partial: " + ", ".join(parts)
            summary_cell.fill = PARTIAL_FILL
        else:
            summary_cell.value = "GAP"
            summary_cell.fill = GAP_FILL

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    for i in range(len(all_col_ids)):
        ws.column_dimensions[get_column_letter(fixed_cols + 1 + i)].width = 6
    ws.column_dimensions[get_column_letter(summary_col)].width = 22

    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=fixed_cols + 1)


def write_coverage_sheet(ws, artifacts, reverse):
    """Write a coverage summary sheet grouped by domain."""
    ws.cell(row=1, column=1, value="Traceability Coverage Summary").font = Font(bold=True, name="Arial", size=14)
    ws.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}").font = Font(name="Arial", size=10, italic=True)

    headers = ["ID", "Title", "Status", "Domain", "Has User Story", "Has Task", "Has Test", "Coverage"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=i + 1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    row_ids = sorted([aid for aid in artifacts if get_prefix(aid) in ROW_TYPES],
                     key=lambda x: (artifacts[x].get("domain", ""), x))

    stats = {"covered": 0, "partial": 0, "gap": 0}

    for row_idx, rid in enumerate(row_ids):
        r = 5 + row_idx
        info = artifacts[rid]

        # Check what links to this requirement
        has_us = False
        has_task = False
        has_test = False

        # Check forward links on other artifacts
        for aid, ainfo in artifacts.items():
            prefix = get_prefix(aid)
            links = ainfo.get("_links", {})
            if prefix == "US" and rid in links.get("delivers", []):
                has_us = True
            if prefix == "TASK" and rid in links.get("implements", []):
                has_task = True
            if prefix == "TEST" and rid in links.get("verifies", []):
                has_test = True

        # Also check legacy reverse fields on the requirement itself
        own_links = info.get("_links", {})
        if own_links.get("delivered_by"):
            has_us = True
        if own_links.get("implemented_by"):
            has_task = True
        if own_links.get("verified_by"):
            has_test = True

        ws.cell(row=r, column=1, value=rid).font = BODY_FONT
        ws.cell(row=r, column=2, value=info["title"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=info["status"]).font = BODY_FONT
        ws.cell(row=r, column=4, value=info.get("domain", "")).font = BODY_FONT

        for col, val in [(5, has_us), (6, has_task), (7, has_test)]:
            cell = ws.cell(row=r, column=col, value="Yes" if val else "No")
            cell.font = BODY_FONT
            cell.fill = OK_FILL if val else GAP_FILL
            cell.alignment = Alignment(horizontal="center")

        cov_cell = ws.cell(row=r, column=8)
        cov_cell.font = Font(bold=True, name="Arial", size=10)
        if has_us and has_test:
            cov_cell.value = "Covered"
            cov_cell.fill = OK_FILL
            stats["covered"] += 1
        elif has_us or has_task or has_test:
            cov_cell.value = "Partial"
            cov_cell.fill = PARTIAL_FILL
            stats["partial"] += 1
        else:
            cov_cell.value = "GAP"
            cov_cell.fill = GAP_FILL
            stats["gap"] += 1

        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Summary stats
    total = len(row_ids)
    summary_row = 5 + len(row_ids) + 2
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, name="Arial", size=12)
    ws.cell(row=summary_row + 1, column=1, value="Total Requirements").font = BODY_FONT
    ws.cell(row=summary_row + 1, column=2, value=total).font = BODY_FONT
    ws.cell(row=summary_row + 2, column=1, value="Fully Covered").font = BODY_FONT
    ws.cell(row=summary_row + 2, column=2, value=stats["covered"]).font = BODY_FONT
    ws.cell(row=summary_row + 2, column=2).fill = OK_FILL
    ws.cell(row=summary_row + 3, column=1, value="Partially Covered").font = BODY_FONT
    ws.cell(row=summary_row + 3, column=2, value=stats["partial"]).font = BODY_FONT
    ws.cell(row=summary_row + 3, column=2).fill = PARTIAL_FILL
    ws.cell(row=summary_row + 4, column=1, value="Gaps (no links)").font = BODY_FONT
    ws.cell(row=summary_row + 4, column=2, value=stats["gap"]).font = BODY_FONT
    ws.cell(row=summary_row + 4, column=2).fill = GAP_FILL

    if total > 0:
        ws.cell(row=summary_row + 5, column=1, value="Coverage %").font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=summary_row + 5, column=2,
                value=f"{(stats['covered'] / total) * 100:.1f}%").font = Font(bold=True, name="Arial", size=10)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14


def write_orphans_sheet(ws, artifacts):
    """Sheet listing artifacts that nothing references."""
    all_referenced = set()
    for aid, info in artifacts.items():
        for targets in info["_links"].values():
            all_referenced.update(targets)

    skip_prefixes = {"VISION", "META", "GLOSS", "CODEMAP", "PERSONA"}
    orphans = [aid for aid in sorted(artifacts)
               if aid not in all_referenced and get_prefix(aid) not in skip_prefixes]

    ws.cell(row=1, column=1, value="Orphan Artifacts").font = Font(bold=True, name="Arial", size=14)
    ws.cell(row=2, column=1, value="Artifacts not referenced by any other artifact").font = Font(italic=True, name="Arial", size=10)

    headers = ["ID", "Title", "Status", "Type"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=4, column=i + 1, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    for idx, oid in enumerate(orphans):
        r = 5 + idx
        info = artifacts.get(oid, {})
        ws.cell(row=r, column=1, value=oid).font = BODY_FONT
        ws.cell(row=r, column=2, value=info.get("title", "")).font = BODY_FONT
        ws.cell(row=r, column=3, value=info.get("status", "")).font = BODY_FONT
        ws.cell(row=r, column=4, value=get_prefix(oid)).font = BODY_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10


def main():
    parser = argparse.ArgumentParser(description="Generate traceability matrix (XLSX)")
    parser.add_argument("--path", default=".", help="Root of requirements vault")
    parser.add_argument("--output", default=None, help="Output XLSX file")
    args = parser.parse_args()

    root = Path(args.path)
    artifacts = build_artifacts(root)
    reverse = compute_reverse_links(artifacts)

    print(f"Found {len(artifacts)} artifacts")

    wb = Workbook()

    # Sheet 1: Cross-reference matrix
    ws_matrix = wb.active
    ws_matrix.title = "Traceability Matrix"
    write_matrix_sheet(ws_matrix, artifacts, reverse)

    # Sheet 2: Coverage summary
    ws_coverage = wb.create_sheet("Coverage Summary")
    write_coverage_sheet(ws_coverage, artifacts, reverse)

    # Sheet 3: Orphans
    ws_orphans = wb.create_sheet("Orphan Artifacts")
    write_orphans_sheet(ws_orphans, artifacts)

    output_path = Path(args.output) if args.output else root / "05-quality" / "traceability" / "TRACEABILITY-MATRIX.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Traceability matrix saved: {output_path}")


if __name__ == "__main__":
    main()
